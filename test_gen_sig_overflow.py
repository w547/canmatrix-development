import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))
import canmatrix
import canmatrix.formats
import openpyxl

# Step 1: Import DBC
db_path = 'test_min_gen_sig.dbc'
dbs = canmatrix.formats.loadp(db_path)

for db_name, db in dbs.items():
    for frame in db.frames:
        for sig in frame.signals:
            gsv = sig.attributes.get("GenSigStartValue")
            print(f"Signal: {sig.name}")
            print(f"  GenSigStartValue attr: {repr(gsv)} (type: {type(gsv).__name__})")
            print(f"  initial_value: {sig.initial_value} (type: {type(sig.initial_value).__name__})")
            print(f"  factor: {sig.factor}, offset: {sig.offset}")

# Step 2: Export to XLSX
xlsx_path = 'test_min_gen_sig.xlsx'
canmatrix.formats.dumpp(dbs, xlsx_path)
print(f"\nExported to {xlsx_path}")

# Step 3: Check Excel contents
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active
headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers}")

if 'GenSigStartValue' in headers:
    idx = headers.index('GenSigStartValue')
    for row in ws.iter_rows(min_row=2, values_only=True):
        sig_name = row[headers.index('Signal Name')]
        gsv_val = row[idx]
        print(f"\n  Signal Name: {sig_name}")
        print(f"  GenSigStartValue cell value: {repr(gsv_val)} (type: {type(gsv_val).__name__})")

# Cleanup
for f in [xlsx_path]:
    if os.path.exists(f):
        os.remove(f)