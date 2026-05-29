import sys
import os
import io
import decimal

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

import canmatrix
from canmatrix.formats import dbc, xls, xlsx, xls_common

test_dbc_content = '''VERSION ""

NS_ :
    NS_DESC_
    CM_
    BA_DEF_
    BA_
    VAL_
    BA_DEF_DEF_
    SIG_VALTYPE_

BS_:

BU_: Tester

BO_ 862 TestFrame: 8 Tester
 SG_ TestSignal_Minus2 : 39|32@0+ (1,0) [0|4.29497E+009] "" Tester
 SG_ TestSignal_Minus2147483648 : 39|32@0+ (1,0) [0|4.29497E+009] "" Tester

BA_DEF_ SG_ "GenSigStartValue" FLOAT 0 100000000000;

BA_ "GenSigStartValue" SG_ 862 TestSignal_Minus2 -2;
BA_ "GenSigStartValue" SG_ 862 TestSignal_Minus2147483648 -2147483648;

VAL_ 862 TestSignal_Minus2 -1 "Invalid" ;
VAL_ 862 TestSignal_Minus2147483648 -1 "Invalid" ;
'''

print("=== Test DBC Content ===")
print(test_dbc_content)

dbc_file = io.BytesIO(test_dbc_content.encode('utf-8'))
db = dbc.load(dbc_file)

print("\n=== After DBC Load ===")
for frame in db.frames:
    for sig in frame.signals:
        print(f"Signal: {sig.name}")
        print(f"  attributes: {sig.attributes}")
        print(f"  initial_value: {sig.initial_value} (type: {type(sig.initial_value).__name__})")
        print(f"  factor: {sig.factor}, offset: {sig.offset}")
        print(f"  GenSigStartValue attr: '{sig.attributes.get('GenSigStartValue', 'N/A')}'")

print("\n=== XLS Common - get_signal ===")
for frame in db.frames:
    for sig in frame.signals:
        front, back = xls_common.get_signal(db, frame, sig, "msbreverse")
        print(f"\nSignal: {sig.name}")
        print(f"  front_array: {front}")
        print(f"  back_array: {back}")
        gen_sig_idx = front.index(sig.attributes.get("GenSigStartValue", "")) if sig.attributes.get("GenSigStartValue", "") in front else -1
        print(f"  GenSigStartValue in front_array: index {gen_sig_idx}")

print("\n=== Test XLS Export ===")
xls_out = io.BytesIO()
xls.dump(db, xls_out)
xls_out.seek(0)
print(f"XLS output size: {len(xls_out.getvalue())} bytes")

import xlrd
wb = xlrd.open_workbook(file_contents=xls_out.read())
sh = wb.sheet_by_index(0)
print(f"Sheet rows: {sh.nrows}, cols: {sh.ncols}")
print("Headers:", [sh.cell(0, c).value for c in range(sh.ncols)])
for r in range(1, sh.nrows):
    row_data = [sh.cell(r, c).value for c in range(sh.ncols)]
    print(f"Row {r}: {row_data}")

print("\n=== Test XLSX Export ===")
xlsx_out = io.BytesIO()
xlsx.dump(db, xlsx_out)
xlsx_out.seek(0)

import openpyxl
wb2 = openpyxl.load_workbook(xlsx_out)
ws = wb2.active
print(f"Sheet rows: {ws.max_row}, cols: {ws.max_column}")
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("Headers:", headers)
gen_sig_col = headers.index('GenSigStartValue') + 1 if 'GenSigStartValue' in headers else -1
sig_default_col = headers.index('Signal Default') + 1 if 'Signal Default' in headers else -1
print(f"GenSigStartValue col: {gen_sig_col}, Signal Default col: {sig_default_col}")
for r in range(2, ws.max_row + 1):
    row_data = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    print(f"Row {r}: {row_data}")
    if gen_sig_col > 0:
        cell_val = ws.cell(r, gen_sig_col).value
        print(f"  GenSigStartValue cell: '{cell_val}' (type: {type(cell_val).__name__})")
    if sig_default_col > 0:
        cell_val = ws.cell(r, sig_default_col).value
        print(f"  Signal Default cell: '{cell_val}' (type: {type(cell_val).__name__})")

print("\n=== Test DBC Re-export ===")
dbc_out = io.BytesIO()
dbc.dump(db, dbc_out)
dbc_out.seek(0)
dbc_output = dbc_out.read().decode('utf-8')
print(dbc_output)

print("\n=== Test DBC Re-import ===")
dbc_out.seek(0)
db2 = dbc.load(dbc_out)
for frame in db2.frames:
    for sig in frame.signals:
        print(f"Signal: {sig.name}")
        print(f"  attributes: {sig.attributes}")
        print(f"  initial_value: {sig.initial_value}")
        print(f"  GenSigStartValue attr: '{sig.attributes.get('GenSigStartValue', 'N/A')}'")
