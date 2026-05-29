import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))
import canmatrix
from canmatrix.convert import convert
import openpyxl

# Test: DBC with non-trivial factor/offset
gen_val = '-2147483648'
dbc_content = f'''VERSION ""

NS_ : 
	NS_DESC_
	CM_
	BA_DEF_
	BA_
	VAL_
	CAT_DEF_
	CAT_
	BA_DEF_DEF_
	EV_DATA_
	ENVVAR_DATA_
	SGTYPE_
	SGTYPE_VAL_
	BA_DEF_SGTYPE_
	BA_SGTYPE_
	SIG_TYPE_REF_
	VAL_TABLE_
	SIG_GROUP_
	SIG_VALTYPE_
	SIGTYPE_VALTYPE_
	BO_TX_BU_

BS_:

BU_: ECU1

BO_ 100 TestFrame: 8 ECU1
 SG_ Sig1 : 0|32@1- (0.01,-40) [{gen_val}|2147483647] "" ECU1
 SG_ Sig2 : 0|32@1- (1,0) [{gen_val}|2147483647] "" ECU1
 SG_ Sig3 : 0|32@1+ (1,0) [0|4294967295] "" ECU1

BA_DEF_ SG_ "GenSigStartValue" INT -2147483648 2147483647;

BA_ "GenSigStartValue" SG_ 100 Sig1 {gen_val};
BA_ "GenSigStartValue" SG_ 100 Sig2 {gen_val};
BA_ "GenSigStartValue" SG_ 100 Sig3 -2;

BA_DEF_DEF_ "GenSigStartValue" 0;
'''

dbc_path = 'test_convert.dbc'
xlsx_path = 'test_convert.xlsx'

with open(dbc_path, 'w') as f:
    f.write(dbc_content)

# Use the convert() function (same as Flask API)
convert(dbc_path, xlsx_path)

# Read back and check
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active
headers = [cell.value for cell in ws[1]]

gsv_idx = headers.index('GenSigStartValue')
sname_idx = headers.index('Signal Name')
sdefault_idx = headers.index('Signal Default')

for row in ws.iter_rows(min_row=2, values_only=True):
    print(f"Signal Name: {row[sname_idx]}")
    print(f"  GenSigStartValue: {repr(row[gsv_idx])} (type: {type(row[gsv_idx]).__name__})")
    print(f"  Signal Default:   {repr(row[sdefault_idx])} (type: {type(row[sdefault_idx]).__name__})")

# Cleanup
for f in [dbc_path, xlsx_path]:
    try:
        os.remove(f)
    except:
        pass