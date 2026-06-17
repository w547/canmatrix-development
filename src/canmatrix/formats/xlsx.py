# -*- coding: utf-8 -*-
# Copyright (c) 2013, Eduard Broecker
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without modification, are permitted provided that
# the following conditions are met:
#
#    Redistributions of source code must retain the above copyright notice, this list of conditions and the
#    following disclaimer.
#    Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the
#    following disclaimer in the documentation and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED
# WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A
# PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY
# DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO,
# PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR
# OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH
# DAMAGE.

#
# this script exports xls-files from a canmatrix-object
# xls-files are the can-matrix-definitions displayed in Excel

import logging
import typing
from builtins import *

import openpyxl
import openpyxl.utils

import canmatrix
import canmatrix.formats.xls_common
import decimal
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# Font Size : 8pt * 20 = 160
# font = 'font: name Arial Narrow, height 160'
font = 'font: name Verdana, height 160'

sty_header = 0
sty_norm = 0
sty_first_frame = 0
sty_white = 0

sty_green = 0
sty_green_first_frame = 0
sty_sender = 0
sty_sender_first_frame = 0
sty_sender_green = 0
sty_sender_green_first_frame = 0


def write_ecu_matrix(ecu_list, signal, frame, worksheet, row, col, first_frame):
    # type: (typing.Sequence[str], typing.Optional[canmatrix.Signal], canmatrix.Frame, xlsxwriter.workbook.Worksheet, int, int, xlsxwriter.workbook.Format) -> int
    # first-frame - style with borders:
    if first_frame == sty_first_frame:
        norm = sty_first_frame
        sender = sty_sender_first_frame
        norm_green = sty_green_first_frame
        sender_green = sty_sender_green_first_frame
    # consecutive-frame - style without borders:
    else:
        norm = sty_norm
        sender = sty_sender
        norm_green = sty_green
        sender_green = sty_sender_green

    # iterate over ECUs:
    for ecu in ecu_list:
        # every second ECU with other style
        if col % 2 == 0:
            loc_style = norm
            loc_style_sender = sender
        # every second ECU with other style
        else:
            loc_style = norm_green
            loc_style_sender = sender_green
        # write "s" "r" "r/s" if signal is sent, received or send and received by ECU
        if signal is not None and ecu in signal.receivers and ecu in frame.transmitters:
            worksheet.cell(row=row+1, column=col+1).value = "r/s"
            worksheet.cell(row=row+1, column=col+1).style = loc_style_sender
        elif signal is not None and ecu in signal.receivers:
            worksheet.cell(row=row+1, column=col+1).value = "r"
            worksheet.cell(row=row+1, column=col+1).style = loc_style
        elif signal is not None:
            # For signal rows: don't fall through to frame receivers,
            # only check frame transmitters (to show "s" for sender ECU)
            if ecu in frame.transmitters:
                worksheet.cell(row=row+1, column=col+1).value = "s"
                worksheet.cell(row=row+1, column=col+1).style = loc_style_sender
            else:
                worksheet.cell(row=row+1, column=col+1).value = ""
                worksheet.cell(row=row+1, column=col+1).style = loc_style
        elif ecu in frame.receivers and ecu in frame.transmitters:
            worksheet.cell(row=row+1, column=col+1).value = "r/s"
            worksheet.cell(row=row+1, column=col+1).style = loc_style_sender
        elif ecu in frame.receivers:
            worksheet.cell(row=row+1, column=col+1).value = "r"
            worksheet.cell(row=row+1, column=col+1).style = loc_style
        elif ecu in frame.transmitters:
            worksheet.cell(row=row+1, column=col+1).value = "s"
            worksheet.cell(row=row+1, column=col+1).style = loc_style_sender
        else:
            worksheet.cell(row=row+1, column=col+1).value = ""
            worksheet.cell(row=row+1, column=col+1).style = loc_style
        col += 1
    # loop over ECUs ends here
    return col


def write_excel_line(worksheet, row, col, row_array, style):
    # type: (openpyxl.workbook.Worksheet, int, int, typing.Sequence[typing.Any], xlsxwriter.workbook.Format) -> int
    for item in row_array:
        worksheet.cell(row=row+1, column=col+1).value = item
        if style != 0:
            worksheet.cell(row=row + 1, column=col + 1).style = style
        col += 1
    return col


def dump(db, filename, **options):
    # type: (canmatrix.CanMatrix, str, **str) -> None
    motorola_bit_format = options.get("xlsMotorolaBitFormat", "msbreverse")
    values_in_seperate_lines = options.get("xlsValuesInSeperateLines", True)
    additional_signal_columns = [x for x in options.get("additionalSignalAttributes", "").split(",") if x]
    additional_frame_columns = [x for x in options.get("additionalFrameAttributes", "").split(",") if x]

    head_top = ['ID', 'Frame Name', 'Signal Name']

    frame_columns = [
        'DLC',
        'frame.comment',
        'Cycle Time [ms]',
        'Launch Type',
        'GenMsgDelayTime',
        'DiagRequest',
        'DiagResponse',
        'DiagState',
        'NmMessage',
        'GenMsgILSupport',
        'GenMsgCycleTimeFast',
        'GenMsgNrOfRepetition',
        'CANFD_BRS',
        'ID-Format']

    signal_columns = [
        'Signal Byte No.',
        'Signal Bit No.',
        'Signal Function',
        'Signal Length [Bit]',
        'Signal Default',
        'GenSigStartValue',
        'GenSigInactiveValue',
        'GenSigSendType',
        'EventCommandSignal',
        'GatewayedSignals',
        'GenSigInvalidValue',
        'GenSigTimeoutValue',
        'Factor',
        'Offset',
        'Signal Not Available',
        'Byteorder']

    head_tail = ['Value', 'Name / Phys. Range', 'Function / Increment Unit']

    PREFIX_COUNT = len(head_top)
    FRAME_ATTR_COUNT = len(frame_columns)
    SIGNAL_ATTR_COUNT = len(signal_columns)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'K-Matrix '
    worksheet.sheet_properties.outlinePr.summaryBelow = False

    global sty_header
    global sty_white
    global sty_first_frame
    global sty_norm
    global sty_green
    global sty_green_first_frame
    global sty_sender
    global sty_sender_first_frame
    global sty_sender_green
    global sty_sender_green_first_frame

    sty_header = NamedStyle(name="sty_header")
    sty_header.font = Font(bold=True, size=8, name='Verdana')
    sty_header.alignment = Alignment(horizontal='center', vertical='center')

    sty_header_frame = NamedStyle(name="sty_header_frame")
    sty_header_frame.font = Font(bold=True, size=8, name='Verdana', color='FFFFFF')
    sty_header_frame.fill = PatternFill(patternType='solid', fgColor='366092')
    sty_header_frame.alignment = Alignment(horizontal='center', vertical='center')

    sty_header_ecu = NamedStyle(name="sty_header_ecu")
    sty_header_ecu.font = Font(bold=True, size=8, name='Verdana', color='FFFFFF')
    sty_header_ecu.fill = PatternFill(patternType='solid', fgColor='76933C')
    sty_header_ecu.alignment = Alignment(horizontal='center', vertical='center')

    sty_header_signal = NamedStyle(name="sty_header_signal")
    sty_header_signal.font = Font(bold=True, size=8, name='Verdana', color='FFFFFF')
    sty_header_signal.fill = PatternFill(patternType='solid', fgColor='E26B0A')
    sty_header_signal.alignment = Alignment(horizontal='center', vertical='center')

    sty_remark = NamedStyle(name="sty_remark")
    sty_remark.font = Font(bold=False, size=8, name='Verdana', color='000000')
    sty_remark.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
    sty_remark.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    sty_first_frame = NamedStyle(name="sty_first_frame")
    sty_first_frame.font = Font(bold=True, size=8, name='Verdana', color='ff000000')
    sty_first_frame.fill = PatternFill(patternType='solid', fgColor='DCE6F1')
    sty_first_frame.border = Border(top=Side(border_style='thin'))
    sty_first_frame.alignment = Alignment(horizontal='center', vertical='center')

    sty_white = NamedStyle(name="sty_white")
    sty_white.font = Font(bold=True, size=8, name='Verdana', color='00ffffff')

    sty_norm = NamedStyle(name="sty_norm")
    sty_norm.font = Font(bold=True, size=8, name='Verdana', color='ff000000')
    sty_norm.alignment = Alignment(horizontal='center', vertical='center')

    sty_green = NamedStyle(name="sty_green")
    sty_green.fill = PatternFill(patternType='solid', fgColor='CCFFCC')
    sty_green.alignment = Alignment(horizontal='center', vertical='center')

    sty_green_first_frame = NamedStyle(name="sty_green_first_frame")
    sty_green_first_frame.fill = PatternFill(patternType='solid', fgColor='DCE6F1')
    sty_green_first_frame.border = Border(top=Side(border_style='thin'))
    sty_green_first_frame.alignment = Alignment(horizontal='center', vertical='center')

    sty_sender = NamedStyle(name="sty_sender")
    sty_sender.fill = PatternFill(patternType='lightGrid', fgColor='C0C0C0')
    sty_sender.alignment = Alignment(horizontal='center', vertical='center')

    sty_sender_first_frame = NamedStyle(name="sty_sender_first_frame")
    sty_sender_first_frame.fill = PatternFill(patternType='solid', fgColor='DCE6F1')
    sty_sender_first_frame.border = Border(top=Side(border_style='thin'))
    sty_sender_first_frame.alignment = Alignment(horizontal='center', vertical='center')

    sty_sender_green = NamedStyle(name="sty_sender_green")
    sty_sender_green.fill = PatternFill(patternType='lightGrid', fgColor='C0C0C0', bgColor='CCFFCC')
    sty_sender_green.alignment = Alignment(horizontal='center', vertical='center')

    sty_sender_green_first_frame = NamedStyle(name="sty_sender_green_first_frame")
    sty_sender_green_first_frame.fill = PatternFill(patternType='solid', fgColor='DCE6F1')
    sty_sender_green_first_frame.border = Border(top=Side(border_style='thin'))
    sty_sender_green_first_frame.alignment = Alignment(horizontal='center', vertical='center')

    ecu_list = [ecu.name for ecu in db.ecus]
    ecu_count = len(ecu_list)

    row_array = head_top + ecu_list + frame_columns + signal_columns + head_tail

    for additional_col in additional_frame_columns:
        row_array.append("frame." + additional_col)

    for additional_col in additional_signal_columns:
        row_array.append("signal." + additional_col)

    total_cols = len(row_array)

    ecu_start = PREFIX_COUNT + 1
    ecu_end = ecu_start + ecu_count - 1
    frame_start = ecu_end + 1
    frame_end = frame_start + FRAME_ATTR_COUNT - 1
    signal_start = frame_end + 1
    signal_end = signal_start + SIGNAL_ATTR_COUNT - 1

    for col_idx in range(1, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter] = ColumnDimension(worksheet, customWidth=True)
        worksheet.column_dimensions[col_letter].width = 2

    for col_idx in range(1, total_cols + 1):
        cell = worksheet.cell(row=2, column=col_idx)
        cell.value = row_array[col_idx - 1]
        if col_idx <= 2:
            cell.style = sty_header_frame
        elif col_idx == 3:
            cell.style = sty_header_signal
        elif ecu_start <= col_idx <= ecu_end:
            cell.style = sty_header_ecu
        elif frame_start <= col_idx <= frame_end:
            cell.style = sty_header_frame
        elif signal_start <= col_idx <= signal_end:
            cell.style = sty_header_signal
        else:
            cell.style = sty_header

    remark_text = "备注：1、蓝色表头为报文及其属性；2、橙色表头为信号及其属性；3、绿色表头为ECU矩阵；4、点击最左侧，表格外的\"1\"，可以收起二级菜单，仅展示【报文行】"
    remark_cell = worksheet.cell(row=1, column=1)
    remark_cell.value = remark_text
    remark_cell.style = sty_remark
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    if db.type == canmatrix.matrix_class.CAN:
        frame_hash = {}
        logger.debug("DEBUG: Length of db.frames is %d", len(db.frames))
        for frame in db.frames:
            if frame.is_complex_multiplexed:
                logger.error("Export complex multiplexers is not supported - frame %s might be uncomplete", frame.name)
            frame_hash[int(frame.arbitration_id.id)] = frame
    else:
        frame_hash = {a.name:a for a in db.frames}

    row = 2

    for idx in sorted(frame_hash.keys()):

        frame = frame_hash[idx]

        sig_hash = {}
        for sig in frame.signals:
            if motorola_bit_format == "msb":
                sig_hash["%03d" % int(sig.get_startbit(bit_numbering=1)) + sig.name] = sig
            elif motorola_bit_format == "msbreverse":
                sig_hash["%03d" % int(sig.get_startbit()) + sig.name] = sig
            else:
                sig_hash["%03d" % int(sig.get_startbit(bit_numbering=1, start_little=True)) + sig.name] = sig

        additional_frame_info = [frame.attribute(additional, default="") for additional in additional_frame_columns]

        frame_info = canmatrix.formats.xls_common.get_frame_info(db, frame)

        write_excel_line(worksheet, row, 0, [frame_info[0], frame_info[1], ""], sty_first_frame)

        col = PREFIX_COUNT
        col = write_ecu_matrix(ecu_list, None, frame, worksheet, row, col, sty_first_frame)

        write_excel_line(worksheet, row, col, frame_info[2:], sty_first_frame)
        col += FRAME_ATTR_COUNT

        write_excel_line(worksheet, row, col, ["" for _ in range(SIGNAL_ATTR_COUNT)], sty_first_frame)
        col += SIGNAL_ATTR_COUNT

        tail_row = ["" for _ in range(len(head_tail))]
        tail_row += additional_frame_info
        tail_row += ["" for _ in additional_signal_columns]
        write_excel_line(worksheet, row, col, tail_row, sty_first_frame)
        row += 1

        if len(sig_hash) == 0:
            continue

        signal_style = sty_norm

        for sig_idx in sorted(sig_hash.keys()):
            sig = sig_hash[sig_idx]

            worksheet.row_dimensions[row+1].outline_level = 1

            (frontRow, back_row) = canmatrix.formats.xls_common.get_signal(db, frame, sig, motorola_bit_format)
            signal_name = frontRow[2]
            signal_data = frontRow[0:2] + frontRow[3:]

            if len(sig.values) > 0 and not values_in_seperate_lines:
                value_style = signal_style
                for val in sorted(sig.values.keys()):
                    write_excel_line(worksheet, row, 0, ["", "", signal_name], signal_style)

                    col = PREFIX_COUNT
                    col = write_ecu_matrix(ecu_list, sig, frame, worksheet, row, col, signal_style)

                    write_excel_line(worksheet, row, col, ["" for _ in range(FRAME_ATTR_COUNT)], signal_style)
                    col += FRAME_ATTR_COUNT

                    write_excel_line(worksheet, row, col, signal_data, signal_style)
                    col += SIGNAL_ATTR_COUNT

                    back_row_copy = list(back_row)
                    back_row_copy += ["" for _ in additional_frame_columns]
                    for item in additional_signal_columns:
                        temp = getattr(sig, item, "")
                        back_row_copy.append(temp)

                    write_excel_line(worksheet, row, col + 2, back_row_copy, signal_style)
                    write_excel_line(worksheet, row, col, [val, sig.values[val]], value_style)

                    row += 1
                    signal_style = sty_norm
                    value_style = sty_norm

            else:
                write_excel_line(worksheet, row, 0, ["", "", signal_name], signal_style)

                col = PREFIX_COUNT
                col = write_ecu_matrix(ecu_list, sig, frame, worksheet, row, col, signal_style)

                write_excel_line(worksheet, row, col, ["" for _ in range(FRAME_ATTR_COUNT)], signal_style)
                col += FRAME_ATTR_COUNT

                write_excel_line(worksheet, row, col, signal_data, signal_style)
                col += SIGNAL_ATTR_COUNT

                if float(sig.min) != 0 or float(sig.max) != 1.0:
                    back_row.insert(0, str(sig.min) + ".." + str(sig.max))
                else:
                    back_row.insert(0, "")
                back_row.insert(0, "")

                back_row += ["" for _ in additional_frame_columns]
                for item in additional_signal_columns:
                    temp = getattr(sig, item, "")
                    back_row.append(temp)

                write_excel_line(worksheet, row, col, back_row, signal_style)
                if len(sig.values) > 0:
                    write_excel_line(worksheet, row, col, ["\n".join(["{}: {}".format(a,b) for (a,b) in sig.values.items()])], signal_style)
                row += 1
                signal_style = sty_norm

    for col_idx in range(1, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_width = 0
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                lines = str(cell.value).split('\n')
                for line in lines:
                    max_width = max(max_width, len(line))
        adjusted_width = min(max_width + 4, 60)
        worksheet.column_dimensions[col_letter].width = max(adjusted_width, 4)

    frame_comment_col = frame_start + 1
    signal_function_col = frame_start + FRAME_ATTR_COUNT + 2
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(frame_comment_col)].width = 30
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(signal_function_col)].width = 30

    for col_idx in range(1, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        is_empty = True
        for row_idx in range(3, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != '':
                is_empty = False
                break
        if is_empty:
            worksheet.column_dimensions[col_letter].hidden = True

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = worksheet['D3']
    workbook.save(filename=filename)


def load(file, **options):
    # type: (typing.BinaryIO, **str) -> canmatrix.CanMatrix

    all_ecu_names = []

    # else use this hack to read xlsx
    motorola_bit_format = options.get("xlsMotorolaBitFormat", "msbreverse")
    workbook = openpyxl.open(file)
    sheet = workbook._sheets[0]
    db = canmatrix.CanMatrix()
    # Defines not imported...

    canmatrix.formats.xls_common.initialize_excel_attribute_defines(db)

    launch_types = []  # type: typing.List[str]
    launch_type_default = None  # type: typing.Optional[str]
    sig_send_types = []  # type: typing.List[str]
    sig_send_type_default = None  # type: typing.Optional[str]

    ecu_start = ecu_end = 0

    column_heads = [sheet.cell(1,a).value for a in range(1, sheet.max_column+1)]

    header_row = 1
    if column_heads and column_heads[0] is not None and '备注' in str(column_heads[0]):
        header_row = 2
        column_heads = [sheet.cell(2, a).value for a in range(1, sheet.max_column + 1)]

    if 'Signal Name' in column_heads and 'DLC' in column_heads:
        ecu_start = column_heads.index('Signal Name') + 1
        ecu_end = column_heads.index('DLC')
    elif 'ID-Format' in column_heads:
        ecu_start = column_heads.index('ID-Format') + 1
        ecu_end = column_heads.index('Value')
    elif 'Byteorder' in column_heads:
        ecu_start = column_heads.index('Byteorder') + 1
        ecu_end = column_heads.index('Value')
    else:
        ecu_start = column_heads.index('Signal Not Available') + 1
        ecu_end = column_heads.index('Value')

    # ECUs:
    for x in range(ecu_start, ecu_end):
        db.add_ecu(canmatrix.Ecu(column_heads[x]))
        all_ecu_names.append(column_heads[x])
    # initialize:
    frame_id = None
    signal_name = ""
    signal_length = 8
    new_frame = None  # type: typing.Optional[canmatrix.Frame]
    new_signal = None  # type: typing.Optional[canmatrix.Signal]

    def get_if_possible(my_row, my_value, default=None):
        if my_value in column_heads and my_row[column_heads.index(my_value)].value is not None:
            return my_row[column_heads.index(my_value)].value
        else:
            return default

    def _is_frame_row(row):
        id_val = row[column_heads.index('ID')].value
        return id_val is not None and str(id_val).strip() != '' and id_val != 'ID'

    def _is_signal_row(row):
        sig_name = get_if_possible(row, 'Signal Name')
        return sig_name is not None and str(sig_name).strip() != '' and str(sig_name).strip() != 'Signal Name'

    for row in sheet.rows:
        if row[0].row < header_row + 1:
            continue
        if not _is_frame_row(row) and not _is_signal_row(row):
            continue

        is_new_format_frame_row = _is_frame_row(row) and not _is_signal_row(row)

        if _is_frame_row(row):
            row_frame_id = row[column_heads.index('ID')].value
            if row_frame_id != frame_id:
                frame_id = row_frame_id
                frame_name = row[column_heads.index('Frame Name')].value
                cycle_time = get_if_possible(row, 'Cycle Time [ms]', '0')
                launch_type = get_if_possible(row, 'Launch Type')
                dlc = int(get_if_possible(row, 'DLC', '8'))

                if frame_id.endswith("xh"):
                    new_frame = canmatrix.Frame(frame_name, canmatrix.ArbitrationId(int(frame_id[:-2], 16), extended=True), size=dlc)
                else:
                    new_frame = canmatrix.Frame(frame_name, arbitration_id=int(frame_id[:-1], 16), size=dlc)

                if 'frame.comment' in column_heads:
                    comment_val = row[column_heads.index('frame.comment')].value
                    if comment_val is not None:
                        new_frame.add_comment(str(comment_val))

                for col_head in column_heads:
                    if col_head is not None and col_head.startswith("frame.") and col_head != "frame.comment":
                        command_str = col_head.replace("frame", "new_frame")
                        command_str += "=" + str(row[column_heads.index(col_head)].value)
                        exec(command_str)

                db.add_frame(new_frame)

                if launch_type is not None:
                    result = canmatrix.formats.xls_common._import_attr_with_default(
                        new_frame, "GenMsgSendType", launch_type,
                        db=db, defines_dict=db.frame_defines, collect_list=launch_types)
                    if result is not None and launch_type_default is None:
                        launch_type_default = result

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "GenMsgDelayTime", get_if_possible(row, 'GenMsgDelayTime'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "DiagRequest", get_if_possible(row, 'DiagRequest'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "DiagResponse", get_if_possible(row, 'DiagResponse'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "DiagState", get_if_possible(row, 'DiagState'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "NmMessage", get_if_possible(row, 'NmMessage'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "GenMsgILSupport", get_if_possible(row, 'GenMsgILSupport'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "GenMsgCycleTimeFast", get_if_possible(row, 'GenMsgCycleTimeFast'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "GenMsgNrOfRepetition", get_if_possible(row, 'GenMsgNrOfRepetition'), db=db, defines_dict=db.frame_defines)

                canmatrix.formats.xls_common._import_attr_with_default(new_frame, "CANFD_BRS", get_if_possible(row, 'CANFD_BRS'), db=db, defines_dict=db.frame_defines)

                id_format = get_if_possible(row, 'ID-Format')
                if id_format is not None and str(id_format).strip() != '':
                    id_format_str = str(id_format).strip()
                    if '_FD' in id_format_str:
                        new_frame.is_fd = True
                    new_frame.add_attribute("VFrameFormat", id_format_str)

                new_frame.cycle_time = cycle_time

            if is_new_format_frame_row:
                for x in range(ecu_start, ecu_end):
                    cell_val = row[x].value
                    if cell_val is not None:
                        cell_str = str(cell_val)
                        if 's' in cell_str:
                            new_frame.add_transmitter(column_heads[x])
                        if 'r' in cell_str:
                            new_frame.add_receiver(column_heads[x])
                signal_name = ""
                continue

        if not _is_signal_row(row):
            continue

        # new signal detected
        if get_if_possible(row, 'Signal Name') != signal_name:
            receiver = []  # type: typing.List[str]
            start_byte = int(get_if_possible(row, 'Signal Byte No.', "0"))
            start_bit = int(get_if_possible(row, 'Signal Bit No.', "0"))
            signal_name = get_if_possible(row, 'Signal Name')
            signal_comment = get_if_possible(row, 'Signal Function')
            signal_length = int(get_if_possible(row, 'Signal Length [Bit]', 0))
            signal_default = get_if_possible(row, 'Signal Default')
            # signal_sna = get_if_possible(row, 'Signal Not Available')
            multiplex = None  # type: typing.Union[str, int, None]
            if signal_comment is not None and signal_comment.startswith('Mode Signal:'):
                multiplex = 'Multiplexor'
                signal_comment = signal_comment[12:]
            elif signal_comment is not None and signal_comment.startswith('Mode '):
                rest = signal_comment[4:].split(':', 1)
                if len(rest) >= 2:
                    multiplex = int(rest[0].strip())
                    signal_comment = rest[1]
                else:
                    multiplex = None

            signal_byte_order = get_if_possible(row, 'Byteorder')
            if signal_byte_order is not None:
                if 'i' in signal_byte_order:
                    is_little_endian = True
                else:
                    is_little_endian = False
            else:
                is_little_endian = True  # Default Intel

            is_signed = False

            if signal_name != "-":
                for ecu_name in all_ecu_names:
                    ecu_sender_receiver = get_if_possible(row, ecu_name)
                    if ecu_sender_receiver is not None:
                        if 's' in ecu_sender_receiver:
                            new_frame.add_transmitter(ecu_name)
                        if 'r' in ecu_sender_receiver:
                            receiver.append(ecu_name)
                new_signal = canmatrix.Signal(signal_name,
                                              start_bit=(start_byte - 1) * 8 + start_bit,
                                              size=signal_length,
                                              is_little_endian=is_little_endian,
                                              is_signed=is_signed,
                                              receivers=receiver,
                                              multiplex=multiplex)
                if not is_little_endian:
                    # motorola
                    if motorola_bit_format == "msb":
                        new_signal.set_startbit(
                            (start_byte - 1) * 8 + start_bit, bitNumbering=1)
                    elif motorola_bit_format == "msbreverse":
                        new_signal.set_startbit((start_byte - 1) * 8 + start_bit)
                    else:  # motorola_bit_format == "lsb"
                        new_signal.set_startbit(
                            (start_byte - 1) * 8 + start_bit,
                            bitNumbering=1,
                            startLittle=True
                        )
                                    
                if signal_name is not None:
                    new_frame.add_signal(new_signal)
                    new_signal.add_comment(signal_comment)
                    if signal_default is not None and signal_default != '':
                        try:
                            new_signal.initial_value = new_signal.float_factory(signal_default)
                        except (ValueError, decimal.InvalidOperation):
                            pass
                    gen_sig_start_value = get_if_possible(row, 'GenSigStartValue')
                    if gen_sig_start_value is not None and str(gen_sig_start_value).strip() != '':
                        stripped, is_default = canmatrix.formats.xls_common._strip_default_mark(str(gen_sig_start_value).strip())
                        if not is_default:
                            new_signal.add_attribute("GenSigStartValue", stripped)

                    gen_sig_inactive_value = get_if_possible(row, 'GenSigInactiveValue')
                    if gen_sig_inactive_value is not None and str(gen_sig_inactive_value).strip() != '':
                        new_signal.add_attribute("GenSigInactiveValue", str(gen_sig_inactive_value).strip())

                    result = canmatrix.formats.xls_common._import_attr_with_default(new_signal, "GenSigSendType", get_if_possible(row, 'GenSigSendType'), db=db, defines_dict=db.signal_defines, collect_list=sig_send_types)
                    if result is not None and sig_send_type_default is None:
                        sig_send_type_default = result

                    event_command_signal = get_if_possible(row, 'EventCommandSignal')
                    if event_command_signal is not None and str(event_command_signal).strip() != '':
                        new_signal.add_attribute("EventCommandSignal", str(event_command_signal).strip())

                    gatewayed_signals = get_if_possible(row, 'GatewayedSignals')
                    if gatewayed_signals is not None and str(gatewayed_signals).strip() != '':
                        new_signal.add_attribute("GatewayedSignals", str(gatewayed_signals).strip())

                    gen_sig_invalid_value = get_if_possible(row, 'GenSigInvalidValue')
                    if gen_sig_invalid_value is not None and str(gen_sig_invalid_value).strip() != '':
                        new_signal.add_attribute("GenSigInvalidValue", str(gen_sig_invalid_value).strip())

                    gen_sig_timeout_value = get_if_possible(row, 'GenSigTimeoutValue')
                    if gen_sig_timeout_value is not None and str(gen_sig_timeout_value).strip() != '':
                        new_signal.add_attribute("GenSigTimeoutValue", str(gen_sig_timeout_value).strip())
                # function = get_if_possible(row, 'Function / Increment Unit')
        value = get_if_possible(row, 'Value')
        if value is not None:
            value = str(value)
        value_name = get_if_possible(row, 'Name / Phys. Range')

        if value_name == 0 or value_name is None:
            value_name = "0"
        elif value_name == 1:
            value_name = "1"
        test = value_name
        # .encode('utf-8')

        factor = get_if_possible(row, 'Function / Increment Unit')
        if factor is not None:
            if not isinstance(factor, str):
                factor = str(factor)
            factor = factor.strip()
            if " " in factor and factor[0].isdigit():
                (factor, unit) = factor.split(" ", 1)
                factor = factor.strip()
                unit = unit.strip()
                new_signal.unit = unit
                new_signal.factor = float(factor)
            else:
                unit = factor.strip()
                new_signal.unit = unit
                new_signal.factor = 1

        # Helper: parse inline value table (multi-line format from dump)
        def _parse_inline_value_table(value_str):
            for line in value_str.strip().split('\n'):
                line = line.strip()
                if ':' in line:
                    key, val = line.split(':', 1)
                    try:
                        new_signal.add_values(int(float(key.strip())), val.strip())
                    except (ValueError, TypeError):
                        pass

        is_multiline_value = value is not None and '\n' in str(value)

        if ".." in test:
            if is_multiline_value:
                _parse_inline_value_table(str(value))
            elif value is not None and ':' in str(value):
                _parse_inline_value_table(str(value))
            (mini, maxi) = test.strip().split("..", 2)
            try:
                new_signal.offset = new_signal.float_factory(mini)
                new_signal.min = new_signal.float_factory(mini)
                new_signal.max = new_signal.float_factory(maxi)
            except ValueError:
                new_signal.offset = 0
                new_signal.min = None
                new_signal.max = None

        elif len(value_name) > 0:
            if is_multiline_value:
                _parse_inline_value_table(str(value))
            elif value is not None and value.strip():
                try:
                    value_int = int(float(value))
                    new_signal.add_values(value_int, value_name)
                except ValueError:
                    pass
            maxi = pow(2, signal_length) - 1
            new_signal.max = float(maxi)
        else:
            if is_multiline_value:
                _parse_inline_value_table(str(value))
            new_signal.offset = 0
            if signal_length > 0:
                new_signal.max = new_signal.float_factory(pow(2, signal_length) - 1)
            else:
                new_signal.max = new_signal.float_factory(1)
            new_signal.min = new_signal.float_factory(0)

        for col_head in column_heads: # todo explain this possibly dangerous code with eval
            if col_head is not None and col_head.startswith("signal."):
                command_str = col_head.replace("signal", "new_signal")
                command_str += "=" + str(row[column_heads.index(col_head)].value)
                exec(command_str)

        explicit_factor = get_if_possible(row, 'Factor')
        if explicit_factor is not None and str(explicit_factor).strip() != '':
            try:
                new_signal.factor = float(str(explicit_factor).strip())
            except (ValueError, decimal.InvalidOperation):
                pass

        explicit_offset = get_if_possible(row, 'Offset')
        if explicit_offset is not None and str(explicit_offset).strip() != '':
            try:
                new_signal.offset = new_signal.float_factory(str(explicit_offset).strip())
            except (ValueError, decimal.InvalidOperation):
                pass



    # dlc-estimation / ensure minimum DLC based on signals (preserves user-set DLC if sufficient)
    for frame in db.frames:
        frame.update_receiver()
        frame.calc_dlc()

    launch_type_enum = "ENUM"
    for launch_type in launch_types:
        if len(launch_type) > 0:
            launch_type_enum += ' "' + launch_type + '",'
    db.add_frame_defines("GenMsgSendType", launch_type_enum[:-1])
    if launch_type_default is not None:
        db.add_define_default("GenMsgSendType", launch_type_default)

    sig_send_type_enum = "ENUM"
    for sig_send_type in sig_send_types:
        if len(sig_send_type) > 0:
            sig_send_type_enum += ' "' + sig_send_type + '",'
    db.add_signal_defines("GenSigSendType", sig_send_type_enum[:-1])
    if sig_send_type_default is not None:
        db.add_define_default("GenSigSendType", sig_send_type_default)

    db.set_fd_type()
    return db
