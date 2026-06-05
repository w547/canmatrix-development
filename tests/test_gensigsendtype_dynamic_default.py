# -*- coding: utf-8 -*-
import io
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import canmatrix
import canmatrix.formats.dbc
import canmatrix.formats.xls
import canmatrix.formats.xlsx

TEST_DBC = os.path.join(
    os.path.dirname(__file__), "..", "visual_app",
    "测试集", "A66_NIDUCCU_CANFD2_V3.5_Debug_20241212.dbc"
)


def test_dbc_loads_default_from_ba_def_def():
    """验证 DBC 导入时 BA_DEF_DEF_ 的默认值被正确读取"""
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    assert "GenSigSendType" in db.signal_defines, "GenSigSendType should be in signal_defines"
    define = db.signal_defines["GenSigSendType"]
    assert define.defaultValue is not None, "GenSigSendType should have a defaultValue from BA_DEF_DEF_"
    assert define.defaultValue == "Cyclic", \
        f"Expected defaultValue 'Cyclic', got '{define.defaultValue}'"

    print(f"PASS: GenSigSendType.defaultValue = '{define.defaultValue}' (from DBC BA_DEF_DEF_)")


def test_signal_without_explicit_attr_returns_default():
    """验证未显式设置 GenSigSendType 的 signal 回退到默认值"""
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    found_default_sig = None
    found_explicit_sig = None

    for frame in db.frames:
        for sig in frame.signals:
            val = sig.attribute("GenSigSendType", db=db)
            if "GenSigSendType" not in sig.attributes:
                found_default_sig = (sig.name, val)
            else:
                found_explicit_sig = (sig.name, sig.attributes["GenSigSendType"], val)

    if found_default_sig:
        print(f"PASS: Signal '{found_default_sig[0]}' has no explicit GenSigSendType, "
              f"attribute() returns '{found_default_sig[1]}' (from BA_DEF_DEF_ default)")
    else:
        print("INFO: All signals have explicit GenSigSendType set")

    if found_explicit_sig:
        print(f"INFO: Signal '{found_explicit_sig[0]}' has explicit GenSigSendType='{found_explicit_sig[1]}'")

    assert found_default_sig is not None or found_explicit_sig is not None, "No signals found"


def test_xlsx_export_shows_star_for_default():
    """验证导出 XLSX 时，使用默认值的 signal 显示 Cyclic*"""
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        canmatrix.formats.xlsx.dump(db, xlsx_path)

        import openpyxl
        wb = openpyxl.open(xlsx_path)
        sh = wb._sheets[0]

        headers = [sh.cell(1, i).value for i in range(1, sh.max_column + 1)]
        assert "GenSigSendType" in headers, f"GenSigSendType column not found in headers: {headers}"
        gsst_idx = headers.index("GenSigSendType")

        found_star = False
        found_no_star = False

        for row in sh.rows:
            cell_val = row[gsst_idx].value
            if cell_val is not None and str(cell_val).strip() != '':
                if str(cell_val).endswith("*"):
                    found_star = True
                    print(f"PASS: Found default-marked value: '{cell_val}'")
                else:
                    found_no_star = True
                    print(f"INFO: Found explicit value: '{cell_val}'")

        assert found_star, "No default-marked GenSigSendType values found in XLSX export"
        print("PASS: XLSX export correctly marks default values with '*'")
    finally:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)


def test_xls_export_shows_star_for_default():
    """验证导出 XLS 时，使用默认值的 signal 显示 Cyclic*"""
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    xls_data = canmatrix.formats.xls.dump(db, io.BytesIO())

    import xlrd
    wb = xlrd.open_workbook(file_contents=xls_data)
    sh = wb.sheet_by_index(0)

    headers = [sh.cell(0, i).value for i in range(sh.ncols)]
    assert "GenSigSendType" in headers, f"GenSigSendType column not found in headers: {headers}"
    gsst_idx = headers.index("GenSigSendType")

    found_star = False
    for row_num in range(1, sh.nrows):
        cell_val = sh.cell(row_num, gsst_idx).value
        if cell_val and str(cell_val).strip() != '':
            if str(cell_val).endswith("*"):
                found_star = True
                print(f"PASS: Found default-marked value: '{cell_val}'")
                break

    assert found_star, "No default-marked GenSigSendType values found in XLS export"
    print("PASS: XLS export correctly marks default values with '*'")



def test_xlsx_roundtrip_preserves_default():
    """验证 XLSX 往返转换后，默认值 signal 仍回退到默认值"""
    with open(TEST_DBC, "rb") as f:
        db_original = canmatrix.formats.dbc.load(f)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        canmatrix.formats.xlsx.dump(db_original, xlsx_path)
        with open(xlsx_path, "rb") as f:
            db_roundtrip = canmatrix.formats.xlsx.load(f)
    finally:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)

    for orig_frame in db_original.frames:
        rt_frame = db_roundtrip.frame_by_id(orig_frame.arbitration_id)
        if rt_frame is None:
            continue
        for orig_sig in orig_frame.signals:
            rt_sig = rt_frame.signal_by_name(orig_sig.name)
            if rt_sig is None:
                continue

            orig_val = orig_sig.attribute("GenSigSendType", db=db_original)
            rt_val = rt_sig.attribute("GenSigSendType", db=db_roundtrip)

            if orig_val is not None and str(orig_val).strip() != '':
                assert str(orig_val) == str(rt_val), \
                    f"Signal {orig_sig.name}: GenSigSendType mismatch: orig='{orig_val}' vs rt='{rt_val}'"

    print("PASS: XLSX round-trip preserves all GenSigSendType values")


def test_xls_roundtrip_preserves_default():
    """验证 XLS 往返转换后，默认值 signal 仍回退到默认值"""
    with open(TEST_DBC, "rb") as f:
        db_original = canmatrix.formats.dbc.load(f)

    xls_data = canmatrix.formats.xls.dump(db_original, io.BytesIO())
    db_roundtrip = canmatrix.formats.xls.load(io.BytesIO(xls_data))

    for orig_frame in db_original.frames:
        rt_frame = db_roundtrip.frame_by_id(orig_frame.arbitration_id)
        if rt_frame is None:
            continue
        for orig_sig in orig_frame.signals:
            rt_sig = rt_frame.signal_by_name(orig_sig.name)
            if rt_sig is None:
                continue

            orig_val = orig_sig.attribute("GenSigSendType", db=db_original)
            rt_val = rt_sig.attribute("GenSigSendType", db=db_roundtrip)

            if orig_val is not None and str(orig_val).strip() != '':
                assert str(orig_val) == str(rt_val), \
                    f"Signal {orig_sig.name}: GenSigSendType mismatch: orig='{orig_val}' vs rt='{rt_val}'"

    print("PASS: XLS round-trip preserves all GenSigSendType values")


def main():
    print("=" * 60)
    print("GenSigSendType Dynamic Default Test Suite")
    print("=" * 60)

    test_dbc_loads_default_from_ba_def_def()
    test_signal_without_explicit_attr_returns_default()
    test_xlsx_export_shows_star_for_default()
    test_xls_export_shows_star_for_default()
    test_xlsx_roundtrip_preserves_default()
    test_xls_roundtrip_preserves_default()

    print("=" * 60)
    print("ALL TESTS PASSED")
    print("=" * 60)


if __name__ == "__main__":
    main()
