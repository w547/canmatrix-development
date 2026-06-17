# -*- coding: utf-8 -*-
import io
import os
import sys
import tempfile

import canmatrix
import canmatrix.formats.dbc
import canmatrix.formats.xls
import canmatrix.formats.xlsx


TEST_DBC = os.path.join(os.path.dirname(__file__), "files", "dbc", "test_genmsg_repetition.dbc")


def load_dbc(path):
    with open(path, "rb") as f:
        db = canmatrix.formats.dbc.load(f)
    return db


def dump_xls(db):
    f = io.BytesIO()
    canmatrix.formats.xls.dump(db, f)
    f.seek(0)
    return f.read()


def load_xls(data):
    f = io.BytesIO(data)
    db = canmatrix.formats.xls.load(f)
    return db


def dump_xlsx(db, path):
    canmatrix.formats.xlsx.dump(db, path)


def load_xlsx(path):
    with open(path, "rb") as f:
        db = canmatrix.formats.xlsx.load(f)
    return db


def test_dbc_uses_genmsgnrofrepetition():
    """Verify DBC with GenMsgNrOfRepetition (German 'Nr.' naming) loads correctly."""
    db = load_dbc(TEST_DBC)

    frame = db.frame_by_id(canmatrix.ArbitrationId(768))
    assert frame is not None
    assert str(frame.attribute("GenMsgNrOfRepetition")) == "3"
    assert str(frame.attribute("GenMsgCycleTimeFast")) == "20"
    print("PASS: DBC with GenMsgNrOfRepetition loads correctly")


def test_xls_export_contains_repetition_value():
    """Verify XLS export contains GenMsgNoOfRepetitions column with value from GenMsgNrOfRepetition."""
    db = load_dbc(TEST_DBC)
    xls_data = dump_xls(db)

    import xlrd
    wb = xlrd.open_workbook(file_contents=xls_data)
    sh = wb.sheet_by_index(0)

    headers = [sh.cell(0, i).value for i in range(sh.ncols)]
    assert "GenMsgNoOfRepetitions" in headers, f"Column 'GenMsgNoOfRepetitions' not found in headers: {headers}"

    gen_msg_no_idx = headers.index("GenMsgNoOfRepetitions")
    gen_msg_ctf_idx = headers.index("GenMsgCycleTimeFast")

    found_repetition = False
    found_ctf = False
    for row_num in range(1, sh.nrows):
        row_data = [sh.cell(row_num, i).value for i in range(sh.ncols)]
        if 768 in row_data or "InteractionFrame" in str(row_data):
            rep_val = sh.cell(row_num, gen_msg_no_idx).value
            ctf_val = sh.cell(row_num, gen_msg_ctf_idx).value
            if rep_val and str(rep_val).strip():
                assert str(rep_val).strip() == "3", f"Expected '3', got '{rep_val}'"
                found_repetition = True
            if ctf_val and str(ctf_val).strip():
                assert str(ctf_val).strip() == "20", f"Expected '20', got '{ctf_val}'"
                found_ctf = True

    assert found_repetition, "GenMsgNoOfRepetitions value not found in XLS export"
    assert found_ctf, "GenMsgCycleTimeFast value not found in XLS export"
    print("PASS: XLS export contains GenMsgNoOfRepetitions value from GenMsgNrOfRepetition")


def test_xlsx_export_contains_repetition_value():
    """Verify XLSX export contains GenMsgNoOfRepetitions column with value from GenMsgNrOfRepetition."""
    db = load_dbc(TEST_DBC)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        dump_xlsx(db, xlsx_path)

        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb.active

        headers = [sheet.cell(1, i).value for i in range(1, sheet.max_column + 1)]
        assert "GenMsgNoOfRepetitions" in headers, f"Column 'GenMsgNoOfRepetitions' not found in headers: {headers}"

        gen_msg_no_idx = headers.index("GenMsgNoOfRepetitions") + 1
        gen_msg_ctf_idx = headers.index("GenMsgCycleTimeFast") + 1

        found_repetition = False
        found_ctf = False
        for row_num in range(2, sheet.max_row + 1):
            id_val = sheet.cell(row_num, 1).value
            if id_val and str(id_val).strip() == "300h":
                rep_val = sheet.cell(row_num, gen_msg_no_idx).value
                ctf_val = sheet.cell(row_num, gen_msg_ctf_idx).value
                if rep_val is not None and str(rep_val).strip():
                    assert str(rep_val).strip() == "3", f"Expected '3', got '{rep_val}'"
                    found_repetition = True
                if ctf_val is not None and str(ctf_val).strip():
                    assert str(ctf_val).strip() == "20", f"Expected '20', got '{ctf_val}'"
                    found_ctf = True

        assert found_repetition, "GenMsgNoOfRepetitions value not found in XLSX export"
        assert found_ctf, "GenMsgCycleTimeFast value not found in XLSX export"
        print("PASS: XLSX export contains GenMsgNoOfRepetitions value from GenMsgNrOfRepetition")
    finally:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)


def test_xls_roundtrip_preserves_repetition():
    """Test DBC (GenMsgNrOfRepetition) -> XLS -> DBC round-trip preserves the attribute."""
    db_original = load_dbc(TEST_DBC)
    xls_data = dump_xls(db_original)
    db_roundtrip = load_xls(xls_data)

    orig_frame = db_original.frame_by_id(canmatrix.ArbitrationId(768))
    rt_frame = db_roundtrip.frame_by_id(canmatrix.ArbitrationId(768))
    assert rt_frame is not None

    orig_val = orig_frame.attribute("GenMsgNrOfRepetition")
    rt_val = rt_frame.attribute("GenMsgNoOfRepetitions")
    assert str(orig_val) == str(rt_val), \
        f"GenMsgRepetition mismatch: orig='{orig_val}' vs rt='{rt_val}'"

    orig_ctf = orig_frame.attribute("GenMsgCycleTimeFast")
    rt_ctf = rt_frame.attribute("GenMsgCycleTimeFast")
    assert str(orig_ctf) == str(rt_ctf), \
        f"GenMsgCycleTimeFast mismatch: orig='{orig_ctf}' vs rt='{rt_ctf}'"

    print("PASS: XLS round-trip preserves GenMsgNrOfRepetition attribute")


def test_xlsx_roundtrip_preserves_repetition():
    """Test DBC (GenMsgNrOfRepetition) -> XLSX -> DBC round-trip preserves the attribute."""
    db_original = load_dbc(TEST_DBC)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        dump_xlsx(db_original, xlsx_path)
        db_roundtrip = load_xlsx(xlsx_path)

        orig_frame = db_original.frame_by_id(canmatrix.ArbitrationId(768))
        rt_frame = db_roundtrip.frame_by_id(canmatrix.ArbitrationId(768))
        assert rt_frame is not None

        orig_val = orig_frame.attribute("GenMsgNrOfRepetition")
        rt_val = rt_frame.attribute("GenMsgNoOfRepetitions")
        assert str(orig_val) == str(rt_val), \
            f"GenMsgRepetition mismatch: orig='{orig_val}' vs rt='{rt_val}'"

        orig_ctf = orig_frame.attribute("GenMsgCycleTimeFast")
        rt_ctf = rt_frame.attribute("GenMsgCycleTimeFast")
        assert str(orig_ctf) == str(rt_ctf), \
            f"GenMsgCycleTimeFast mismatch: orig='{orig_ctf}' vs rt='{rt_ctf}'"

        print("PASS: XLSX round-trip preserves GenMsgNrOfRepetition attribute")
    finally:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)


def test_all_three_naming_variants():
    """Verify all three naming variants (No, Nr, Nr without s) are recognized."""
    db = canmatrix.CanMatrix()
    canmatrix.formats.xls_common.initialize_excel_attribute_defines(db)

    frame = canmatrix.Frame("TestFrame", arbitration_id=100, size=8)
    db.add_frame(frame)

    frame.add_attribute("GenMsgCycleTimeFast", "42")
    frame.add_attribute("GenMsgNoOfRepetitions", "5")
    frame.add_attribute("GenMsgNrOfRepetitions", "6")
    frame.add_attribute("GenMsgNrOfRepetition", "7")

    import canmatrix.formats.xls_common as xc

    result = xc._get_attr_with_fallback(
        frame,
        ["GenMsgNoOfRepetitions", "GenMsgNrOfRepetitions", "GenMsgNrOfRepetition"],
        db, db.frame_defines
    )
    assert result == "5", f"Should return first match '5', got '{result}'"

    frame2 = canmatrix.Frame("TestFrame2", arbitration_id=200, size=8)
    db.add_frame(frame2)
    frame2.add_attribute("GenMsgNrOfRepetition", "7")

    result2 = xc._get_attr_with_fallback(
        frame2,
        ["GenMsgNoOfRepetitions", "GenMsgNrOfRepetitions", "GenMsgNrOfRepetition"],
        db, db.frame_defines
    )
    assert result2 == "7", f"Should return '7', got '{result2}'"

    print("PASS: All three naming variants recognized correctly")


def main():
    print("=" * 60)
    print("GenMsg Repetition Attribute Round-Trip Test Suite")
    print("=" * 60)

    test_dbc_uses_genmsgnrofrepetition()
    test_xls_export_contains_repetition_value()
    test_xlsx_export_contains_repetition_value()
    test_xls_roundtrip_preserves_repetition()
    test_xlsx_roundtrip_preserves_repetition()
    test_all_three_naming_variants()

    print("=" * 60)
    print("ALL TESTS PASSED")
    print("=" * 60)


if __name__ == "__main__":
    main()
