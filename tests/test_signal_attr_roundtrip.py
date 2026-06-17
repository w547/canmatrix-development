# -*- coding: utf-8 -*-
import io
import os
import sys
import tempfile

import canmatrix
import canmatrix.formats.dbc
import canmatrix.formats.xls
import canmatrix.formats.xlsx


TEST_DBC = os.path.join(os.path.dirname(__file__), "files", "dbc", "test_signal_attributes.dbc")

SIGNAL_ATTR_NAMES = [
    "GenSigStartValue",
    "GenSigInactiveValue",
    "GenSigSendType",
    "EventCommandSignal",
    "GatewayedSignals",
    "GenSigInvalidValue",
    "GenSigTimeoutValue",
]


def load_dbc(path):
    with open(path, "rb") as f:
        db = canmatrix.formats.dbc.load(f)
    return db


def dump_xls(db):
    f = io.BytesIO()
    canmatrix.formats.xls.dump(db, f)
    return f.getvalue()


def load_xls(data):
    f = io.BytesIO(data)
    db = canmatrix.formats.xls.load(f)
    return db


def dump_xlsx(db, path):
    canmatrix.formats.xlsx.dump(db, path)


def load_xlsx(path):
    with open(path, "rb") as f:
        db = canmatrix.formats.xlsx.load(f)
    return db


def get_signal_attrs(signal):
    return {name: signal.attribute(name) for name in SIGNAL_ATTR_NAMES}


def signal_attr_roundtrip_check(db_original, db_roundtrip):
    for orig_frame in db_original.frames:
        rt_frame = db_roundtrip.frame_by_id(orig_frame.arbitration_id)
        assert rt_frame is not None, f"Frame {orig_frame.name} not found after round-trip"

        for orig_sig in orig_frame.signals:
            rt_sig = rt_frame.signal_by_name(orig_sig.name)
            assert rt_sig is not None, f"Signal {orig_sig.name} in frame {orig_frame.name} not found after round-trip"

            orig_attrs = get_signal_attrs(orig_sig)
            rt_attrs = get_signal_attrs(rt_sig)

            for attr_name, orig_val in orig_attrs.items():
                rt_val = rt_attrs[attr_name]
                if orig_val is not None and str(orig_val).strip() != '':
                    if rt_val is not None:
                        assert str(orig_val) == str(rt_val), \
                            f"Signal {orig_sig.name}: attr '{attr_name}' mismatch: orig='{orig_val}' vs rt='{rt_val}'"
                    else:
                        raise AssertionError(
                            f"Signal {orig_sig.name}: attr '{attr_name}' LOST in round-trip: orig='{orig_val}'"
                        )


def test_xls_roundtrip():
    db_original = load_dbc(TEST_DBC)
    xls_data = dump_xls(db_original)
    db_roundtrip = load_xls(xls_data)
    signal_attr_roundtrip_check(db_original, db_roundtrip)
    print("PASS: XLS round-trip test - all signal attributes preserved")


def test_xlsx_roundtrip():
    db_original = load_dbc(TEST_DBC)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        dump_xlsx(db_original, xlsx_path)
        db_roundtrip = load_xlsx(xlsx_path)
    finally:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)

    signal_attr_roundtrip_check(db_original, db_roundtrip)
    print("PASS: XLSX round-trip test - all signal attributes preserved")


def test_expected_values():
    db = load_dbc(TEST_DBC)

    frame1 = db.frame_by_id(canmatrix.ArbitrationId(256))
    assert frame1 is not None

    sig1 = frame1.signal_by_name("Sig1")
    assert sig1 is not None
    assert str(sig1.attribute("GenSigStartValue")) == "100"
    assert str(sig1.attribute("GenSigInactiveValue")) == "0"
    assert str(sig1.attribute("GenSigSendType")) == "Cyclic"
    assert str(sig1.attribute("EventCommandSignal")) == "Yes"
    assert str(sig1.attribute("GatewayedSignals")) == "SigA,SigB"
    assert str(sig1.attribute("GenSigInvalidValue")) == "255"
    assert str(sig1.attribute("GenSigTimeoutValue")) == "500"

    sig2 = frame1.signal_by_name("Sig2")
    assert sig2 is not None
    assert str(sig2.attribute("GenSigStartValue")) == "50"
    assert str(sig2.attribute("GenSigInactiveValue")) == "0xFF"
    assert str(sig2.attribute("GenSigSendType")) == "OnChange"
    assert str(sig2.attribute("EventCommandSignal")) == "No"
    assert str(sig2.attribute("GenSigInvalidValue")) == "0"
    assert str(sig2.attribute("GenSigTimeoutValue")) == "1000"

    frame2 = db.frame_by_id(canmatrix.ArbitrationId(512))
    assert frame2 is not None

    sig3 = frame2.signal_by_name("Sig3")
    assert sig3 is not None
    assert str(sig3.attribute("GenSigStartValue")) == "200"
    assert str(sig3.attribute("GenSigInactiveValue")) == "0xFFFF"
    assert str(sig3.attribute("GenSigSendType")) == "IfActive"
    assert str(sig3.attribute("EventCommandSignal")) == "Yes"
    assert str(sig3.attribute("GatewayedSignals")) == "SigX"
    assert str(sig3.attribute("GenSigInvalidValue")) == "65535"
    assert str(sig3.attribute("GenSigTimeoutValue")) == "2000"

    sig4 = frame2.signal_by_name("Sig4")
    assert sig4 is not None
    assert str(sig4.attribute("GenSigStartValue")) == "0"
    assert str(sig4.attribute("GenSigInactiveValue")) == "0"
    assert str(sig4.attribute("GenSigSendType")) == "NoSigSendType"
    assert str(sig4.attribute("EventCommandSignal")) == "No"
    assert str(sig4.attribute("GenSigInvalidValue")) == "0"
    assert str(sig4.attribute("GenSigTimeoutValue")) == "3000"

    print("PASS: expected values test - all specific signal attributes verified")


def test_xls_export_contains_columns():
    db = load_dbc(TEST_DBC)
    xls_data = dump_xls(db)

    import xlrd
    wb = xlrd.open_workbook(file_contents=xls_data)
    sh = wb.sheet_by_index(0)

    headers = [sh.cell(0, i).value for i in range(sh.ncols)]

    expected_headers = [
        "GenSigStartValue",
        "GenSigInactiveValue",
        "GenSigSendType",
        "EventCommandSignal",
        "GatewayedSignals",
        "GenSigInvalidValue",
        "GenSigTimeoutValue",
    ]
    for h in expected_headers:
        assert h in headers, f"Column '{h}' not found in XLS headers: {headers}"

    found_sig1_attrs = {h: False for h in expected_headers}

    for row_num in range(1, sh.nrows):
        row_data = [sh.cell(row_num, i).value for i in range(sh.ncols)]
        if "Sig1" in str(row_data):
            for h in expected_headers:
                idx = headers.index(h)
                if sh.cell(row_num, idx).value:
                    found_sig1_attrs[h] = True

    for h in expected_headers:
        assert found_sig1_attrs[h], f"Signal attribute '{h}' value not found for Sig1 in XLS data rows"

    print("PASS: XLS column headers and signal attribute values verified")


def test_xlsx_export_contains_columns():
    db = load_dbc(TEST_DBC)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        dump_xlsx(db, xlsx_path)

        import openpyxl
        wb = openpyxl.open(xlsx_path)
        sh = wb._sheets[0]

        headers = [sh.cell(2, i).value for i in range(1, sh.max_column + 1)]

        expected_headers = [
            "GenSigStartValue",
            "GenSigInactiveValue",
            "GenSigSendType",
            "EventCommandSignal",
            "GatewayedSignals",
            "GenSigInvalidValue",
            "GenSigTimeoutValue",
        ]
        for h in expected_headers:
            assert h in headers, f"Column '{h}' not found in XLSX headers: {headers}"

        found_sig1_attrs = {h: False for h in expected_headers}

        for row in sh.rows:
            row_data = [cell.value for cell in row]
            if "Sig1" in str(row_data):
                for h in expected_headers:
                    idx = headers.index(h)
                    if row[idx].value:
                        found_sig1_attrs[h] = True

        for h in expected_headers:
            assert found_sig1_attrs[h], f"Signal attribute '{h}' value not found for Sig1 in XLSX data rows"

        print("PASS: XLSX column headers and signal attribute values verified")
    finally:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)


def test_dbc_to_dbc_preserves_attrs():
    db_original = load_dbc(TEST_DBC)

    f = io.BytesIO()
    canmatrix.formats.dbc.dump(db_original, f)
    f.seek(0)
    db_roundtrip = canmatrix.formats.dbc.load(f)

    signal_attr_roundtrip_check(db_original, db_roundtrip)
    print("PASS: DBC->DBC round-trip test - all signal attributes preserved")


def main():
    print("=" * 60)
    print("Signal Attribute Round-Trip Test Suite")
    print("=" * 60)

    test_expected_values()
    test_dbc_to_dbc_preserves_attrs()
    test_xls_export_contains_columns()
    test_xlsx_export_contains_columns()
    test_xls_roundtrip()
    test_xlsx_roundtrip()

    print("=" * 60)
    print("ALL TESTS PASSED")
    print("=" * 60)


if __name__ == "__main__":
    main()
