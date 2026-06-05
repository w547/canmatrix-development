# -*- coding: utf-8 -*-
"""Test that large signal min/max values survive DBC -> Excel -> DBC round-trip without precision loss."""

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import canmatrix
import canmatrix.formats
from canmatrix.utils import FloatFactory


def _create_db_with_large_signals():
    db = canmatrix.CanMatrix()
    frame = canmatrix.Frame("TestFrame", arbitration_id=canmatrix.ArbitrationId(0x100), size=8)
    db.add_frame(frame)

    test_signals = [
        ("MCUR_SwVerExt_Debug",   0, 32, 0, 4294967295),
        ("Sig_MaxU16",           32, 16, 0, 65535),
        ("Sig_MaxU8",            48,  8, 0, 255),
        ("Sig_LargeMinMax",      56,  8, 200, 250),
    ]

    for name, start_bit, size, min_val, max_val in test_signals:
        sig = canmatrix.Signal(name, start_bit=start_bit, size=size)
        sig.min = FloatFactory.get_float(min_val)
        sig.max = FloatFactory.get_float(max_val)
        sig.factor = FloatFactory.get_float(1.0)
        sig.offset = FloatFactory.get_float(0)
        frame.add_signal(sig)

    db.add_ecu(canmatrix.Ecu("TestECU"))
    return db, test_signals


class TestLargeValueRoundtrip(unittest.TestCase):

    def test_dbc_to_xlsx_to_dbc_roundtrip(self):
        db, test_signals = _create_db_with_large_signals()

        tmpdir = tempfile.mkdtemp()
        try:
            dbc1 = os.path.join(tmpdir, "test.dbc")
            xlsx = os.path.join(tmpdir, "test.xlsx")
            dbc2 = os.path.join(tmpdir, "test2.dbc")

            with open(dbc1, "wb") as f:
                canmatrix.formats.dump(db, f, export_type="dbc", dbcExportEncoding="iso-8859-1")
            self.assertTrue(os.path.getsize(dbc1) > 0, "DBC file empty")

            dbs1 = canmatrix.formats.loadp(dbc1)
            db1 = list(dbs1.values())[0]

            with open(xlsx, "wb") as f:
                canmatrix.formats.dump(db1, f, export_type="xlsx")
            self.assertTrue(os.path.getsize(xlsx) > 0, "XLSX file empty")

            dbs2 = canmatrix.formats.loadp(xlsx)
            db2 = list(dbs2.values())[0]

            with open(dbc2, "wb") as f:
                canmatrix.formats.dump(db2, f, export_type="dbc", dbcExportEncoding="iso-8859-1")
            self.assertTrue(os.path.getsize(dbc2) > 0, "Final DBC file empty")

            dbs3 = canmatrix.formats.loadp(dbc2)
            db3 = list(dbs3.values())[0]

            for name, _, _, exp_min, exp_max in test_signals:
                sig3 = db3.frame_by_name('TestFrame').signal_by_name(name)
                self.assertIsNotNone(sig3, "Signal %s not found after roundtrip" % name)
                self.assertEqual(
                    int(sig3.min), exp_min,
                    "Signal %s: min mismatch: expected %d, got %s" % (name, exp_min, sig3.min)
                )
                self.assertEqual(
                    int(sig3.max), exp_max,
                    "Signal %s: max mismatch: expected %d, got %s" % (name, exp_max, sig3.max)
                )
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)

    def test_xlsx_no_sci_notation_in_any_cell(self):
        db, _ = _create_db_with_large_signals()

        tmpdir = tempfile.mkdtemp()
        try:
            xlsx = os.path.join(tmpdir, "test.xlsx")
            with open(xlsx, "wb") as f:
                canmatrix.formats.dump(db, f, export_type="xlsx")
            self.assertTrue(os.path.getsize(xlsx) > 0, "XLSX file empty")

            import openpyxl
            wb = openpyxl.open(xlsx)
            sheet = wb._sheets[0]

            range_cells_found = 0
            for row in range(2, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row, col)
                    if cell.value and ".." in str(cell.value):
                        range_cells_found += 1
                        self.assertIsInstance(
                            cell.value, str,
                            "Cell %s should be string, got %s: %s" % (cell.coordinate, type(cell.value).__name__, cell.value)
                        )
                        val_str = str(cell.value)
                        self.assertNotIn(
                            "e+", val_str.lower(),
                            "Cell %s contains scientific notation: %s" % (cell.coordinate, cell.value)
                        )
                        self.assertNotIn(
                            "e-", val_str.lower(),
                            "Cell %s contains scientific notation: %s" % (cell.coordinate, cell.value)
                        )
            self.assertGreater(range_cells_found, 0, "No range cells (containing '..') found in xlsx")
            wb.close()
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
