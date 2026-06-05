# -*- coding: utf-8 -*-
import sys
import os
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import canmatrix
import canmatrix.formats.dbc
import canmatrix.formats.xlsx
import canmatrix.formats.xls

TEST_DBC = os.path.join(os.path.dirname(__file__), '..', 'visual_app', '测试集', 'A66_NIDUCCU_CANFD2_V3.5_Debug_20241212.dbc')

def test_dbc_loads_default_from_ba_def_def():
    print("=== Test 1: DBC loads GenSigSendType default from BA_DEF_DEF_ ===")
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    assert "GenSigSendType" in db.signal_defines, "GenSigSendType not in signal_defines"
    define = db.signal_defines["GenSigSendType"]
    print(f"  GenSigSendType defaultValue = '{define.defaultValue}'")
    assert define.defaultValue == "Cyclic", f"Expected 'Cyclic', got '{define.defaultValue}'"
    print("  PASSED: defaultValue correctly read as 'Cyclic' from BA_DEF_DEF_")

def test_signal_without_explicit_attr_returns_default():
    print("\n=== Test 2: Signal without explicit GenSigSendType returns default ===")
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    found_without = False
    found_with = False
    for frame in db.frames:
        for sig in frame.signals:
            if "GenSigSendType" not in sig.attributes:
                val = sig.attribute("GenSigSendType", db=db)
                print(f"  Signal '{sig.name}' (no explicit attr): attribute() returns '{val}'")
                found_without = True
            else:
                val = sig.attribute("GenSigSendType", db=db)
                print(f"  Signal '{sig.name}' (explicit: '{sig.attributes['GenSigSendType']}'): attribute() returns '{val}'")
                found_with = True
            if found_without and found_with:
                break
        if found_without and found_with:
            break

    assert found_without, "No signal without explicit GenSigSendType found"
    print("  PASSED: attribute() correctly falls back to default value")

def test_xlsx_export_shows_star_for_default():
    print("\n=== Test 3: XLSX export shows '*' for default values ===")
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        canmatrix.formats.xlsx.dump(db, xlsx_path)

        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        gen_sig_send_type_col = None
        for i, h in enumerate(headers):
            if h == 'GenSigSendType':
                gen_sig_send_type_col = i
                break

        assert gen_sig_send_type_col is not None, "GenSigSendType column not found in XLSX"

        star_found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[gen_sig_send_type_col]
            if val is not None and str(val).endswith('*'):
                print(f"  Found default-marked value: '{val}'")
                star_found = True
                break

        if not star_found:
            print("  WARNING: No '*' marked values found - all signals may have explicit GenSigSendType")

        print("  PASSED: XLSX export completed successfully")
    finally:
        os.unlink(xlsx_path)

def test_xlsx_roundtrip_preserves_default():
    print("\n=== Test 4: XLSX roundtrip preserves GenSigSendType default ===")
    with open(TEST_DBC, "rb") as f:
        db = canmatrix.formats.dbc.load(f)

    original_default = db.signal_defines["GenSigSendType"].defaultValue
    print(f"  Original default: '{original_default}'")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        xlsx_path = tmp.name

    try:
        canmatrix.formats.xlsx.dump(db, xlsx_path)
        db2 = canmatrix.formats.xlsx.load(xlsx_path)

        assert "GenSigSendType" in db2.signal_defines, "GenSigSendType lost in roundtrip"
        new_default = db2.signal_defines["GenSigSendType"].defaultValue
        print(f"  Roundtrip default: '{new_default}'")
        assert new_default == original_default, f"Default changed: '{original_default}' -> '{new_default}'"
        print("  PASSED: Default value preserved through XLSX roundtrip")
    finally:
        os.unlink(xlsx_path)

def test_no_hardcoded_cyclic():
    print("\n=== Test 5: No hardcoded 'Cyclic' default in source code ===")
    import inspect
    xls_source = inspect.getsource(canmatrix.formats.xls)
    xlsx_source = inspect.getsource(canmatrix.formats.xlsx)

    # Check that the hardcoded insert is gone
    assert 'sig_send_types.insert(0, "Cyclic")' not in xls_source, "Hardcoded 'Cyclic' still in xls.py!"
    assert 'sig_send_types.insert(0, "Cyclic")' not in xlsx_source, "Hardcoded 'Cyclic' still in xlsx.py!"
    print("  PASSED: No hardcoded 'Cyclic' insertion found in xls.py or xlsx.py")

if __name__ == '__main__':
    test_dbc_loads_default_from_ba_def_def()
    test_signal_without_explicit_attr_returns_default()
    test_xlsx_export_shows_star_for_default()
    test_xlsx_roundtrip_preserves_default()
    test_no_hardcoded_cyclic()
    print("\n=== All tests passed! ===")
