import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))
import canmatrix
import canmatrix.formats
import openpyxl

# Test: DBC -> XLSX, then read XLSX back and check GenSigStartValue column
gen_val = '-2147483648'
dbc_content = f'''VERSION ""

NS_ : 
	NS_DESC_
	CM_
	BA_DEF_
	BA_
	VAL_
	CAT_DEF_
	CAT_
	BA_DEF_DEF_
	EV_DATA_
	ENVVAR_DATA_
	SGTYPE_
	SGTYPE_VAL_
	BA_DEF_SGTYPE_
	BA_SGTYPE_
	SIG_TYPE_REF_
	VAL_TABLE_
	SIG_GROUP_
	SIG_VALTYPE_
	SIGTYPE_VALTYPE_
	BO_TX_BU_

BS_:

BU_:

BO_ 100 TestFrame: 8 Vector__XXX
 SG_ TestSignal : 0|32@1- (1,0) [{gen_val}|2147483647] "" Vector__XXX

BA_DEF_ SG_ "GenSigStartValue" INT {gen_val} 2147483647;

BA_ "GenSigStartValue" SG_ 100 TestSignal {gen_val};

BA_DEF_DEF_ "GenSigStartValue" 0;
'''

dbc_path = 'test_roundtrip.dbc'
with open(dbc_path, 'w') as f:
    f.write(dbc_content)

dbs = canmatrix.formats.loadp(dbc_path)

xlsx_path = 'test_roundtrip.xlsx'
canmatrix.formats.dumpp(dbs, xlsx_path)

# Read back
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active
headers = [cell.value for cell in ws[1]]

gsv_idx = headers.index('GenSigStartValue')
sname_idx = headers.index('Signal Name')
sdefault_idx = headers.index('Signal Default')

for row in ws.iter_rows(min_row=2, values_only=True):
    print(f"Signal Name: {row[sname_idx]}")
    print(f"  GenSigStartValue: {repr(row[gsv_idx])} (type: {type(row[gsv_idx]).__name__})")
    print(f"  Signal Default: {repr(row[sdefault_idx])} (type: {type(row[sdefault_idx]).__name__})")

# Now read back via canmatrix import
dbs2 = canmatrix.formats.loadp(xlsx_path)
for db_name, db in dbs2.items():
    for frame in db.frames:
        for sig in frame.signals:
            gsv = sig.attributes.get("GenSigStartValue")
            print(f"\nAfter re-import:")
            print(f"  GenSigStartValue attr: {repr(gsv)} (type: {type(gsv).__name__})")
            print(f"  initial_value: {repr(sig.initial_value)}")

# Cleanup
for f in [dbc_path, xlsx_path]:
    try:
        os.remove(f)
    except:
        pass