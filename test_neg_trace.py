import sys
import os
import io

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from canmatrix.formats import dbc, xls, xlsx
from canmatrix import CanMatrix

# Test DBC with INT 0 0 define and negative GenSigStartValue
test_dbc = '''VERSION "test"

NS_ :

BS_:

BU_: ECU1

BO_ 862 TestFrame: 8 ECU1
 SG_ TestSig1 : 0|32@0+ (1,0) [0|4294970000] "" Tester
 SG_ TestSig2 : 32|32@0+ (1,0) [0|4294970000] "" Tester

BA_DEF_ SG_ "GenSigStartValue" INT 0 0;
BA_DEF_DEF_ "GenSigStartValue" 0;

BA_ "GenSigStartValue" SG_ 862 TestSig1 -2;
BA_ "GenSigStartValue" SG_ 862 TestSig2 -2147483648;
'''

print("=" * 60)
print("INPUT DBC:")
print(test_dbc)

# Test 1: DBC import
print("=" * 60)
print("TEST 1: Import DBC")
db = dbc.load(io.BytesIO(test_dbc.encode('utf-8')))

for frame in db.frames:
    for sig in frame.signals:
        gsv = sig.attributes.get('GenSigStartValue', 'N/A')
        print(f"  {sig.name}: GenSigStartValue='{gsv}', initial_value={sig.initial_value}")
        print(f"    factor={sig.factor}, offset={sig.offset}, is_signed={sig.is_signed}")

if 'GenSigStartValue' in db.signal_defines:
    sd = db.signal_defines['GenSigStartValue']
    print(f"  Define: type={sd.type}, min={sd.min}, max={sd.max}, default={sd.defaultValue}")

# Test 2: DBC export (roundtrip)
print("\n" + "=" * 60)
print("TEST 2: Export to DBC")
out = io.BytesIO()
dbc.dump(db, out)
out.seek(0)
output = out.read().decode('utf-8')
for line in output.split('\n'):
    if 'GenSigStartValue' in line:
        print(f"  {line.strip()}")

# Test 3: Re-import
print("\n" + "=" * 60)
print("TEST 3: Re-import from exported DBC")
out.seek(0)
db2 = dbc.load(out)
for frame in db2.frames:
    for sig in frame.signals:
        gsv = sig.attributes.get('GenSigStartValue', 'N/A')
        print(f"  {sig.name}: GenSigStartValue='{gsv}', initial_value={sig.initial_value}")

# Test 4: Export to XLSX
print("\n" + "=" * 60)
print("TEST 4: Export to XLSX")
xlsx_out = io.BytesIO()
xlsx.dump(db, xlsx_out)
xlsx_out.seek(0)

# Read XLSX cells directly
from openpyxl import load_workbook
wb = load_workbook(xlsx_out)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"  Sheet: {sheet_name}")
    for row in ws.iter_rows(min_row=1, values_only=False):
        row_values = [str(cell.value) for cell in row]
        if any('TestSig' in str(v) for v in row_values):
            # Find GenSigStartValue column
            for cell in row:
                print(f"    col={cell.column}, value={cell.value} (type={type(cell.value).__name__})")
            print(f"    Row: {row_values}")

# Test 5: Re-import from XLSX
print("\n" + "=" * 60)
print("TEST 5: Re-import from XLSX")
xlsx_out.seek(0)
db3 = xlsx.load(xlsx_out)
for frame in db3.frames:
    for sig in frame.signals:
        gsv = sig.attributes.get('GenSigStartValue', 'N/A')
        print(f"  {sig.name}: GenSigStartValue='{gsv}', initial_value={sig.initial_value}")

# Test 6: Export re-imported data to DBC
print("\n" + "=" * 60)
print("TEST 6: Export from XLSX-reimported data to DBC")
dbc_out2 = io.BytesIO()
dbc.dump(db3, dbc_out2)
dbc_out2.seek(0)
output2 = dbc_out2.read().decode('utf-8')
for line in output2.split('\n'):
    if 'GenSigStartValue' in line:
        print(f"  {line.strip()}")

# Test 7: Same but with XLS
print("\n" + "=" * 60)
print("TEST 7: Export to XLS")
xls_out = io.BytesIO()
xls.dump(db, xls_out)
xls_out.seek(0)

# Re-import from XLS
db4 = xls.load(xls_out)
for frame in db4.frames:
    for sig in frame.signals:
        gsv = sig.attributes.get('GenSigStartValue', 'N/A')
        print(f"  {sig.name}: GenSigStartValue='{gsv}', initial_value={sig.initial_value}")

# Export to DBC
dbc_out3 = io.BytesIO()
dbc.dump(db4, dbc_out3)
dbc_out3.seek(0)
output3 = dbc_out3.read().decode('utf-8')
for line in output3.split('\n'):
    if 'GenSigStartValue' in line:
        print(f"  {line.strip()}")
