import sys
import os
import io
import decimal

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

import canmatrix
from canmatrix.formats import dbc, xls, xlsx, xls_common
from canmatrix import convert

test_cases = [
    {
        "name": "FLOAT define, explicit negative value (-2)",
        "content": '''VERSION ""

NS_ :
    NS_DESC_
    CM_
    BA_DEF_
    BA_
    VAL_
    BA_DEF_DEF_
    SIG_VALTYPE_

BS_:

BU_: Tester

BO_ 862 TestFrame: 8 Tester
 SG_ Sig_Neg2 : 39|32@0+ (1,0) [0|4.29497E+009] "" Tester

BA_DEF_ SG_ "GenSigStartValue" FLOAT 0 100000000000;

BA_ "GenSigStartValue" SG_ 862 Sig_Neg2 -2;

VAL_ 862 Sig_Neg2 -1 "Invalid" ;
'''
    },
    {
        "name": "FLOAT define, explicit negative value (-2147483648)",
        "content": '''VERSION ""

NS_ :
    NS_DESC_
    CM_
    BA_DEF_
    BA_
    VAL_
    BA_DEF_DEF_
    SIG_VALTYPE_

BS_:

BU_: Tester

BO_ 862 TestFrame: 8 Tester
 SG_ Sig_NegBig : 39|32@0+ (1,0) [0|4.29497E+009] "" Tester

BA_DEF_ SG_ "GenSigStartValue" FLOAT 0 100000000000;

BA_ "GenSigStartValue" SG_ 862 Sig_NegBig -2147483648;

VAL_ 862 Sig_NegBig -1 "Invalid" ;
'''
    },
    {
        "name": "INT 0 0 define, unsigned value (4294967294) - mimicking 新建文本文档.txt",
        "content": '''VERSION ""

NS_ :
    NS_DESC_
    CM_
    BA_DEF_
    BA_
    VAL_
    BA_DEF_DEF_
    SIG_VALTYPE_

BS_:

BU_: Tester

BO_ 862 TestFrame: 8 Tester
 SG_ Sig_Unsigned : 39|32@0+ (1,0) [0|4.29497E+009] "" Tester

BA_DEF_ SG_ "GenSigStartValue" INT 0 0;
BA_DEF_DEF_ "GenSigStartValue" 0;

BA_ "GenSigStartValue" SG_ 862 Sig_Unsigned 4294967294;

VAL_ 862 Sig_Unsigned -1 "Invalid" ;
'''
    },
    {
        "name": "FLOAT define, large unsigned (4294967294) in DBC",
        "content": '''VERSION ""

NS_ :
    NS_DESC_
    CM_
    BA_DEF_
    BA_
    VAL_
    BA_DEF_DEF_
    SIG_VALTYPE_

BS_:

BU_: Tester

BO_ 862 TestFrame: 8 Tester
 SG_ Sig_Unsigned2 : 39|32@0+ (1,0) [0|4.29497E+009] "" Tester

BA_DEF_ SG_ "GenSigStartValue" FLOAT 0 100000000000;

BA_ "GenSigStartValue" SG_ 862 Sig_Unsigned2 4294967294;

VAL_ 862 Sig_Unsigned2 -1 "Invalid" ;
'''
    },
]

for tc in test_cases:
    print(f"\n{'='*60}")
    print(f"Test: {tc['name']}")
    print(f"{'='*60}")

    dbc_file = io.BytesIO(tc['content'].encode('utf-8'))
    db = dbc.load(dbc_file)

    for frame in db.frames:
        for sig in frame.signals:
            print(f"  Signal: {sig.name}")
            print(f"    attributes: {sig.attributes}")
            print(f"    initial_value: {sig.initial_value} (type: {type(sig.initial_value).__name__})")
            gsv = sig.attributes.get('GenSigStartValue', 'N/A')
            print(f"    GenSigStartValue attr: '{gsv}' (type: {type(gsv).__name__})")

    xls_out = io.BytesIO()
    xls.dump(db, xls_out)
    xls_out.seek(0)

    import xlrd
    wb = xlrd.open_workbook(file_contents=xls_out.read())
    sh = wb.sheet_by_index(0)
    headers = [sh.cell(0, c).value for c in range(sh.ncols)]
    gsv_col = headers.index('GenSigStartValue') if 'GenSigStartValue' in headers else -1
    sd_col = headers.index('Signal Default') if 'Signal Default' in headers else -1
    for r in range(1, sh.nrows):
        if gsv_col >= 0:
            print(f"  XLS GenSigStartValue: '{sh.cell(r, gsv_col).value}' (type: {type(sh.cell(r, gsv_col).value).__name__})")
        if sd_col >= 0:
            print(f"  XLS Signal Default: '{sh.cell(r, sd_col).value}' (type: {type(sh.cell(r, sd_col).value).__name__})")

    xlsx_out = io.BytesIO()
    xlsx.dump(db, xlsx_out)
    xlsx_out.seek(0)

    import openpyxl
    wb2 = openpyxl.load_workbook(xlsx_out)
    ws = wb2.active
    headers2 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    gsv_col2 = headers2.index('GenSigStartValue') + 1 if 'GenSigStartValue' in headers2 else -1
    sd_col2 = headers2.index('Signal Default') + 1 if 'Signal Default' in headers2 else -1
    for r in range(2, ws.max_row + 1):
        if gsv_col2 > 0:
            cv = ws.cell(r, gsv_col2).value
            print(f"  XLSX GenSigStartValue: '{cv}' (type: {type(cv).__name__})")
        if sd_col2 > 0:
            cv = ws.cell(r, sd_col2).value
            print(f"  XLSX Signal Default: '{cv}' (type: {type(cv).__name__})")

    dbc_out = io.BytesIO()
    dbc.dump(db, dbc_out)
    dbc_out.seek(0)
    dbc_output = dbc_out.read().decode('utf-8')
    
    for line in dbc_output.split('\n'):
        if 'GenSigStartValue' in line:
            print(f"  DBC export: {line.strip()}")

    dbc_out.seek(0)
    db2 = dbc.load(dbc_out)
    for frame in db2.frames:
        for sig in frame.signals:
            gsv = sig.attributes.get('GenSigStartValue', 'N/A')
            print(f"  Re-import GenSigStartValue: '{gsv}', initial_value: {sig.initial_value}")

print("\n\n" + "="*60)
print("Test: Web app convert path (DBC -> XLSX)")
print("="*60)

dbc_file = io.BytesIO(test_cases[0]['content'].encode('utf-8'))
db = dbc.load(dbc_file)

xlsx_out = io.BytesIO()
convert.convert(db, xlsx_out, 'xlsx')
xlsx_out.seek(0)

import openpyxl
wb2 = openpyxl.load_workbook(xlsx_out)
ws = wb2.active
headers2 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
gsv_col2 = headers2.index('GenSigStartValue') + 1 if 'GenSigStartValue' in headers2 else -1
sd_col2 = headers2.index('Signal Default') + 1 if 'Signal Default' in headers2 else -1
for r in range(2, ws.max_row + 1):
    if gsv_col2 > 0:
        cv = ws.cell(r, gsv_col2).value
        print(f"  XLSX GenSigStartValue: '{cv}' (type: {type(cv).__name__})")
    if sd_col2 > 0:
        cv = ws.cell(r, sd_col2).value
        print(f"  XLSX Signal Default: '{cv}' (type: {type(cv).__name__})")
